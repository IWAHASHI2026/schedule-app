import csv
import io
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer, PageBreak, Paragraph
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from sqlalchemy.orm import Session
from database import Schedule, ShiftAssignment, Employee, JobType, ShiftRequest
import calendar
import os


JOB_TYPE_COLORS = {
    "職人": "FF6B6B",
    "サブ職人": "4DABF7",
    "lkデータ": "51CF66",
    "uv/cpデータ": "CC5DE8",
    "手紙": "F59F00",
    "その他": "FFD43B",
}


def _get_schedule_data(db: Session, month: str):
    schedule = (
        db.query(Schedule)
        .filter(Schedule.target_month == month)
        .order_by(Schedule.id.desc())
        .first()
    )
    if not schedule:
        raise ValueError("No schedule found for the specified month")

    employees = db.query(Employee).order_by(Employee.sort_order).all()
    job_types = db.query(JobType).order_by(JobType.sort_order).all()
    jt_map = {jt.id: jt.name for jt in job_types}

    assignments = (
        db.query(ShiftAssignment)
        .filter(ShiftAssignment.schedule_id == schedule.id)
        .order_by(ShiftAssignment.employee_id, ShiftAssignment.date)
        .all()
    )

    year, mon = map(int, month.split("-"))
    days_in_month = calendar.monthrange(year, mon)[1]
    dates = [date(year, mon, d) for d in range(1, days_in_month + 1)]

    # 希望休の日付セットを構築
    requests = db.query(ShiftRequest).filter(ShiftRequest.target_month == month).all()
    requested_off: set[tuple[int, date]] = set()
    for req in requests:
        for detail in req.details:
            requested_off.add((req.employee_id, detail.date))

    # Build matrix: emp_id -> date -> job_type_name
    matrix: dict[int, dict[date, str]] = {}
    # Build summary: job_type_id -> date -> headcount
    summary: dict[int, dict[date, float]] = {jt.id: {d: 0.0 for d in dates} for jt in job_types}
    daily_totals: dict[date, float] = {d: 0.0 for d in dates}

    for a in assignments:
        if a.employee_id not in matrix:
            matrix[a.employee_id] = {}
        if a.job_type_id and a.work_type != "off":
            name = jt_map.get(a.job_type_id, "")
            if a.work_type == "morning_half":
                name += "(午前)"
            elif a.work_type == "afternoon_half":
                name += "(午後)"
            matrix[a.employee_id][a.date] = name
            # Accumulate summary
            summary[a.job_type_id][a.date] += a.headcount_value
            daily_totals[a.date] += a.headcount_value
        else:
            if (a.employee_id, a.date) in requested_off:
                matrix[a.employee_id][a.date] = "希休"
            else:
                matrix[a.employee_id][a.date] = "調休"

    return employees, dates, matrix, job_types, summary, daily_totals


def _split_dates(dates: list[date]) -> tuple[list[date], list[date]]:
    """Split dates into first half (day 1-15) and second half (day 16-end)."""
    first_half = [d for d in dates if d.day <= 15]
    second_half = [d for d in dates if d.day > 15]
    return first_half, second_half


def _fmt_val(val: float) -> str:
    """Format a summary value: integer if whole, one decimal otherwise, empty if 0."""
    if val == 0:
        return ""
    if val == int(val):
        return str(int(val))
    return str(val)


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def generate_csv(db: Session, month: str) -> str:
    employees, dates, matrix, job_types, summary, daily_totals = _get_schedule_data(db, month)
    first_half, second_half = _split_dates(dates)

    output = io.StringIO()
    writer = csv.writer(output)

    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]

    for date_slice in [first_half, second_half]:
        # Header
        header = ["スタッフ"] + [f"{d.day}({weekday_names[d.weekday()]})" for d in date_slice]
        writer.writerow(header)

        # Employee rows
        for emp in employees:
            row = [emp.name]
            for d in date_slice:
                row.append(matrix.get(emp.id, {}).get(d, ""))
            writer.writerow(row)

        # Summary rows per job type
        for jt in job_types:
            row = [jt.name]
            for d in date_slice:
                row.append(_fmt_val(summary[jt.id][d]))
            writer.writerow(row)

        # Daily total row
        total_row = ["合計"]
        for d in date_slice:
            total_row.append(_fmt_val(daily_totals[d]))
        writer.writerow(total_row)

        # Blank separator between halves
        writer.writerow([])

    return output.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def generate_excel(db: Session, month: str) -> bytes:
    employees, dates, matrix, job_types, summary, daily_totals = _get_schedule_data(db, month)
    first_half, second_half = _split_dates(dates)

    wb = Workbook()
    ws = wb.active
    ws.title = f"シフト表 {month}"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]
    current_row = 1

    for date_slice in [first_half, second_half]:
        # Header row
        hdr_cell = ws.cell(row=current_row, column=1, value="スタッフ")
        hdr_cell.font = Font(bold=True, size=8)
        hdr_cell.border = thin_border
        ws.column_dimensions['A'].width = 12

        for col_idx, d in enumerate(date_slice, start=2):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = f"{d.day}\n{weekday_names[d.weekday()]}"
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 6
            if d.weekday() >= 5:
                cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        current_row += 1

        # Employee data rows
        for emp in employees:
            ws.cell(row=current_row, column=1, value=emp.name).border = thin_border
            ws.cell(row=current_row, column=1).font = Font(size=8)
            for col_idx, d in enumerate(date_slice, start=2):
                cell = ws.cell(row=current_row, column=col_idx)
                val = matrix.get(emp.id, {}).get(d, "")
                cell.value = val
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                cell.font = Font(size=8)

                if d.weekday() >= 5:
                    cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")
                elif val == "希休":
                    cell.fill = PatternFill(start_color="E9D5FF", fill_type="solid")
                    cell.font = Font(size=8, color="7C3AED")
                elif val == "調休":
                    cell.fill = PatternFill(start_color="E2E8F0", fill_type="solid")
                    cell.font = Font(size=8, color="64748B")
                elif val:
                    for jt_name, color in JOB_TYPE_COLORS.items():
                        if jt_name in val:
                            cell.fill = PatternFill(start_color=color, fill_type="solid")
                            break
            current_row += 1

        # Summary rows per job type
        for jt in job_types:
            name_cell = ws.cell(row=current_row, column=1, value=jt.name)
            name_cell.border = thin_border
            jt_hex = JOB_TYPE_COLORS.get(jt.name)
            name_cell.font = Font(size=8, bold=True, color=jt_hex if jt_hex else "000000")
            name_cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
            for col_idx, d in enumerate(date_slice, start=2):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = _fmt_val(summary[jt.id][d]) or None
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
                cell.font = Font(size=8)
                cell.fill = PatternFill(start_color="F0F0F0", fill_type="solid")
            current_row += 1

        # Daily total row
        total_name = ws.cell(row=current_row, column=1, value="合計")
        total_name.border = thin_border
        total_name.font = Font(size=8, bold=True)
        total_name.fill = PatternFill(start_color="E0E0E0", fill_type="solid")
        for col_idx, d in enumerate(date_slice, start=2):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = _fmt_val(daily_totals[d]) or None
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            cell.font = Font(size=8, bold=True)
            cell.fill = PatternFill(start_color="E0E0E0", fill_type="solid")
        current_row += 1

        # Gap between halves
        current_row += 2

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def generate_pdf(db: Session, month: str) -> bytes:
    employees, dates, matrix, job_types, summary, daily_totals = _get_schedule_data(db, month)
    first_half, second_half = _split_dates(dates)

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    # Try to register a Japanese font
    _try_register_japanese_font()
    font_name = _get_japanese_font_name()
    bold_font_name = _get_japanese_bold_font_name()

    weekday_names = ["月", "火", "水", "木", "金", "土", "日"]

    jt_colors_rgb = {
        "職人": colors.Color(1.0, 0.42, 0.42),
        "サブ職人": colors.Color(0.3, 0.67, 0.97),
        "lkデータ": colors.Color(0.32, 0.81, 0.4),
        "uv/cpデータ": colors.Color(0.80, 0.36, 0.91),
        "手紙": colors.Color(0.96, 0.62, 0.0),
        "その他": colors.Color(1.0, 0.83, 0.23),
    }

    # Year-month title style
    year = int(month.split("-")[0])
    mon = int(month.split("-")[1])
    title_text = f"{year}年{mon}月"
    title_font = bold_font_name or font_name or "Helvetica-Bold"
    title_style = ParagraphStyle(
        "Title",
        fontName=title_font,
        fontSize=14,
        leading=18,
        spaceAfter=4 * mm,
    )

    elements = []

    for date_slice in [first_half, second_half]:
        # Add year-month title
        elements.append(Paragraph(title_text, title_style))

        # Build header
        header = [""] + [f"{d.day}\n{weekday_names[d.weekday()]}" for d in date_slice]
        data = [header]

        # Employee rows
        for emp in employees:
            row = [emp.name]
            for d in date_slice:
                row.append(matrix.get(emp.id, {}).get(d, ""))
            data.append(row)

        num_emp_rows = len(employees)

        # Summary rows per job type
        for jt in job_types:
            row = [jt.name]
            for d in date_slice:
                row.append(_fmt_val(summary[jt.id][d]))
            data.append(row)

        # Daily total row
        total_row = ["合計"]
        for d in date_slice:
            total_row.append(_fmt_val(daily_totals[d]))
        data.append(total_row)

        # Column widths
        name_width = 45
        available_width = landscape(A4)[0] - 16 * mm - name_width
        day_width = max(14, available_width / len(date_slice))
        col_widths = [name_width] + [day_width] * len(date_slice)

        table = Table(data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ]

        # Use Japanese bold font for all cells
        if bold_font_name:
            style_cmds.append(('FONTNAME', (0, 0), (-1, -1), bold_font_name))
        elif font_name:
            style_cmds.append(('FONTNAME', (0, 0), (-1, -1), font_name))
        else:
            style_cmds.append(('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'))

        # Weekend column coloring
        for col_idx, d in enumerate(date_slice):
            if d.weekday() >= 5:
                style_cmds.append(
                    ('BACKGROUND', (col_idx + 1, 0), (col_idx + 1, -1),
                     colors.Color(0.85, 0.85, 0.85))
                )

        # Employee cell coloring (rows 1 .. num_emp_rows)
        for row_idx, emp in enumerate(employees, start=1):
            for col_idx, d in enumerate(date_slice):
                val = matrix.get(emp.id, {}).get(d, "")
                if val == "希休":
                    style_cmds.append(
                        ('BACKGROUND', (col_idx + 1, row_idx), (col_idx + 1, row_idx),
                         colors.Color(0.91, 0.84, 1.0)))
                    style_cmds.append(
                        ('TEXTCOLOR', (col_idx + 1, row_idx), (col_idx + 1, row_idx),
                         colors.Color(0.49, 0.23, 0.93)))
                elif val == "調休":
                    style_cmds.append(
                        ('BACKGROUND', (col_idx + 1, row_idx), (col_idx + 1, row_idx),
                         colors.Color(0.89, 0.91, 0.94)))
                    style_cmds.append(
                        ('TEXTCOLOR', (col_idx + 1, row_idx), (col_idx + 1, row_idx),
                         colors.Color(0.39, 0.45, 0.55)))
                else:
                    for jt_name, color in jt_colors_rgb.items():
                        if jt_name in val:
                            style_cmds.append(
                                ('BACKGROUND', (col_idx + 1, row_idx), (col_idx + 1, row_idx), color))
                            break

        # Summary rows styling (light gray background)
        summary_start = 1 + num_emp_rows
        summary_end = summary_start + len(job_types)  # total row index
        style_cmds.append(
            ('BACKGROUND', (0, summary_start), (-1, summary_end - 1),
             colors.Color(0.94, 0.94, 0.94)))
        # Total row slightly darker
        style_cmds.append(
            ('BACKGROUND', (0, summary_end), (-1, summary_end),
             colors.Color(0.88, 0.88, 0.88)))

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)
        elements.append(PageBreak())

    # Remove trailing PageBreak after last table
    if elements and isinstance(elements[-1], PageBreak):
        elements.pop()

    doc.build(elements)
    return output.getvalue()


def _try_register_japanese_font():
    """Try to register a Japanese font for PDF generation."""
    import glob

    # Bundled font (most reliable - works in all environments)
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    bundled_font = os.path.join(base_dir, "fonts", "NotoSansJP.ttf")

    font_paths = [
        bundled_font,
        # Windows
        "C:/Windows/Fonts/msgothic.ttc",
        "C:/Windows/Fonts/meiryo.ttc",
        # macOS
        "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        # Linux (standard)
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    ]

    # Bold font paths
    bundled_bold = os.path.join(base_dir, "fonts", "NotoSansJP-Bold.ttf")
    bold_font_paths = [
        bundled_bold,
        # Windows
        "C:/Windows/Fonts/msgothic.ttc",
        "C:/Windows/Fonts/meiryob.ttc",
        # macOS
        "/System/Library/Fonts/ヒラギノ角ゴシック W6.ttc",
        # Linux (standard)
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Bold.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
    ]

    # Nix store (Railway/nixpacks): search for Noto CJK fonts dynamically
    nix_fonts = glob.glob("/nix/store/*/share/fonts/noto-cjk/*.ttc")
    font_paths.extend(nix_fonts)
    bold_font_paths.extend(nix_fonts)

    for path in font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("JapaneseFont", path))
                break
            except Exception:
                continue

    # Register bold variant
    for path in bold_font_paths:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("JapaneseFontBold", path))
                return
            except Exception:
                continue


def _get_japanese_font_name() -> str | None:
    """Return registered Japanese font name if available."""
    try:
        pdfmetrics.getFont("JapaneseFont")
        return "JapaneseFont"
    except KeyError:
        return None


def _get_japanese_bold_font_name() -> str | None:
    """Return registered Japanese bold font name if available."""
    try:
        pdfmetrics.getFont("JapaneseFontBold")
        return "JapaneseFontBold"
    except KeyError:
        return _get_japanese_font_name()
